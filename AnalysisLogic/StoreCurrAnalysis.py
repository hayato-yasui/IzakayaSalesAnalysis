import os
import pandas as pd
import numpy as np
import datetime
import seaborn as sns
import matplotlib.pyplot as plt
from multiprocessing import Process
from openpyxl import load_workbook

from Common.Logic.Preprocess import *
from Common.Logic.Postprocess import Postprocess
from Common.Logic.ChartClient import ChartClient
from Common.Setting.StoreCurrAnalysisSetting import *
from Common.Setting.Common.PreprocessSetting import *
from Common.util import Util


# 店舗の現状を把握する為に加工したデータをExcelや図に出力するクラス
class StoreCurrAnalysis:
    def __init__(self):
        self.chart_cli = ChartClient()
        self.preproc_s = PreprocessSetting()
        self.sca_s = StoreCurrAnalysisSetting()
        self.preproc = Preprocess()
        self.sc = SrcConversion()
        self.gu = GroupingUnit()
        self.util = Util()
        self.mmt = MergeMasterTable()
        self.mmt_s = MergeMasterTableSetting()

    # def execute(self,tgt_store):
    def execute(self):
        tgt_store = ['大和乃山賊', '定楽屋', 'うおにく', 'かこい屋', 'くつろぎ屋', 'ご馳走屋名駅店', 'ご馳走屋金山店',
                     '九州乃山賊小倉総本店', '和古屋', '楽屋', '鳥Bouno!', 'ぐるめ屋']
        # tgt_store = ['大和乃山賊', ]
        for s in tgt_store:
            self.sca_s.TGT_STORE = self.preproc_s.TGT_STORE = s
            self.sca_s.OUTPUT_DIR = './data/OUTPUT/' + self.sca_s.TGT_STORE + '/'
            self.preproc_s.DATA_FILES_TO_FETCH = ['売上データ詳細_' + self.preproc_s.TGT_STORE + '_20180401-0630.csv', ]
            self.preproc_s.PROCESSED_DATA_DIR = './data/Input/processed_data/' + self.preproc_s.TGT_STORE + '/'

            self.df_preproc, preproc_csv_file_name = self._preprocess()

            # preproc_csv_file_name = ''
            # self.df_preproc = self.preproc.fetch_csv_and_create_src_df(self.preproc_s.PROCESSED_DATA_DIR
            #                                                            , [preproc_csv_file_name])
            # df_grouped_src = self.df_preproc.groupby(self.cols).mean().reset_index()
            # df_daily = df_grouped_src[self.cols]
            # self.util.df_to_csv(df_daily, self.sca_s.OUTPUT_DIR, '大和乃山賊＿サンプル.csv')

            self._output_store_curr_info(del_old_file=True)

            print(s + " is finish")

    def _preprocess(self):
        df_src = self.preproc.common_proc(self.preproc_s)
        df_src = self.mmt.merge_store(df_src, self.mmt_s.F_PATH_STORE)
        # df_src = self.mmt.merge_weather(df_src, self.mmt_s.DIR_WEATHER, self.preproc_s.TGT_PERIOD_FLOOR,
        #                                        self.preproc_s.TGT_PERIOD_TOP)

        df_src = self.preproc.calc_entering_and_exiting_time(df_src)
        df_src = self.preproc.create_stay_presense(df_src, df_src.loc[0, '営業開始時間'], df_src.loc[0, '営業締め時間'])

        self.preproc.dt_min_round(df_src, '注文時間', 10)
        self.preproc.dt_min_round(df_src, '滞在時間', 20)
        preproc_csv_file_name = self.preproc.create_proc_data_csv(df_src, self.preproc_s.PROCESSED_DATA_DIR,
                                                                  self.preproc_s.TGT_STORE,
                                                                  self.preproc_s.TGT_PERIOD_FLOOR,
                                                                  self.preproc_s.TGT_PERIOD_TOP,
                                                                  memo=self.preproc_s.FILE_MEMO)

        return df_src, preproc_csv_file_name

    def _output_store_curr_info(self, del_old_file=False):
        self.df_grouped_by_bill = self._create_df_grouped_by_bill()
        self.df_set_date_index = self._create_df_set_date_index()

        # self.output_dict = dict()
        # self._monthly_sales()
        # self._daily_cstm_info()
        # self._abc_analysis()
        # self._sheet_occupancy()
        #
        # if del_old_file and os.path.isfile(self.sca_s.OUTPUT_DIR + self.sca_s.OUTPUT_F_EXCEL):
        #     os.remove(self.sca_s.OUTPUT_DIR + self.sca_s.OUTPUT_F_EXCEL)
        # with pd.ExcelWriter(self.sca_s.OUTPUT_DIR + self.sca_s.OUTPUT_F_EXCEL) as writer:
        #     self.util.check_existing_and_create_excel_file(self.sca_s.OUTPUT_DIR + self.sca_s.OUTPUT_F_EXCEL)
        #     writer.book = load_workbook(self.sca_s.OUTPUT_DIR + self.sca_s.OUTPUT_F_EXCEL)
        #     [v_df.to_excel(writer, sheet_name=k, merge_cells=False) for k, v_df in self.output_dict.items()]

        self._plot_moving_avg()

    def _create_df_grouped_by_bill(self):
        return self.df_preproc.groupby(self.gu.BILL).max().reset_index()

    def _create_df_set_date_index(self):
        return self.df_grouped_by_bill.set_index(pd.DatetimeIndex(self.df_grouped_by_bill['H.集計対象営業年月日']))

    def _plot_moving_avg(self):
        df_daily = self.df_set_date_index.groupby(self.df_set_date_index.index). \
            agg(self.sca_s.GROUPING_WAY_DAILY)

        df_daily = self.util.moving_average(df_daily, 'H.伝票金額', 7)
        df_daily = self.util.moving_average(df_daily, 'H.客数（合計）', 7)
        self.chart_cli.plot_axis_is_index(df_daily, needsSave=True,
                                          file_path=self.sca_s.OUTPUT_DIR + '移動平均_' + self.sca_s.TGT_STORE +'.png')

    def _sheet_occupancy(self):
        time_cols = []
        curr_time = self.df_grouped_by_bill.loc[0, '営業開始時間']
        end_time = self.df_grouped_by_bill.loc[0, '営業締め時間']
        sheet_num = int(self.df_grouped_by_bill.loc[0, '席数'])
        while curr_time < end_time:
            if curr_time % 100 == 0:
                curr_time_plus30 = curr_time + 30
            else:
                curr_time_plus30 = curr_time + 70
            time_cols.append(str(curr_time) + '-' + str(curr_time_plus30))
            curr_time = curr_time_plus30

        df_timely_sheet_occupancy = self.df_set_date_index.groupby([
            self.df_set_date_index.index.year.rename('year'),
            self.df_set_date_index.index.month.rename('month'),
            self.df_set_date_index.index.day.rename('day')])[time_cols].sum() / sheet_num

        self.output_dict.update({'座席占有率': df_timely_sheet_occupancy})

    def _monthly_sales(self):
        df_monthly_sales = self.df_set_date_index.groupby([
            self.df_set_date_index.index.year.rename('year'),
            self.df_set_date_index.index.month.rename('month')])['H.伝票金額'].sum()

        self.output_dict.update({'月間売上': df_monthly_sales})

    def _daily_cstm_info(self):
        df_daily_cstm = self.df_set_date_index.groupby(self.df_set_date_index.index). \
            agg(self.sca_s.GROUPING_WAY_DAILY_CSTM)
        self.output_dict.update({'日別客情報': df_daily_cstm})

        # self.chart_cli.create_pie_chart(
        #     df=self.preproc.grouping(self.df_preproc, self.sca_s.GROUPING_KEY_ITEM_CATEGORY2,
        #                              self.sca_s.GROUPING_WAY, self.sca_s.PIE_CHART_SET[0]),
        #     amount_col=self.sca_s.PIE_CHART_SET[1])

        # 時系列カラムをインデックスに指定する必要がある
        # self.chart_cli.time_series_graph(self.df_preproc,
        #                                  amount_cols_li=self.df_preproc[self.sca_s.TIME_SERIES_GRAPH_MONTHLY])
        # self.chart_cli.time_series_graph(self.df_preproc,
        #                                  amount_cols_li=self.df_preproc[self.sca_s.TIME_SERIES_GRAPH_DAYLY])
        #

        # self.chart_cli.plotfig()
        # self.chart_cli.savefig(self.sca_s.OUTPUT_DIR + self.sca_s.FIG_FILE_NAME)
        # self.chart_cli.closefig()

    def _abc_analysis(self):
        [self.sales_and_ratio_by_key(k, True) for k in self.sca_s.ABC_BILL_LEVEL_KEY]
        [self.sales_and_ratio_by_key(k, False) for k in self.sca_s.ABC_NO_BILL_LEVEL_KEY]

    def sales_and_ratio_by_key(self, key_li, bill_level_summary=True):
        df_grouped = self.df_preproc.groupby(key_li)
        df_sales_by_key = df_grouped.agg({'D.価格': np.sum, "H.客数（合計）": np.mean})
        df_sales_by_key['売上比率'] = df_sales_by_key['D.価格'] / int(df_sales_by_key['D.価格'].sum())

        if bill_level_summary:
            df_tmp = self.df_preproc.groupby(self.gu.BILL + key_li).mean().reset_index().set_index(key_li,
                                                                                                   drop=True)
            s_count_by_key = df_tmp.groupby(key_li).size()
        else:
            s_count_by_key = df_grouped.size()
        s_count_by_key.name = 'count_' + '_'.join(key_li)
        s_ratio_by_key = pd.Series(s_count_by_key / s_count_by_key.sum(), name='ratio_' + '_'.join(key_li))

        df_merged = pd.concat([df_sales_by_key, s_count_by_key, s_ratio_by_key], axis=1)
        df_merged = self.preproc.sort_df(df_merged, ['count_' + '_'.join(key_li)], [False])
        df_merged["平均支払額"] = df_merged['D.価格'] / df_merged['count_' + '_'.join(key_li)]

        if key_li in self.sca_s.CALC_PRICE_PER_CSTM:
            df_merged["客単価"] = df_merged["平均支払額"] / df_merged["H.客数（合計）"]
        else:
            df_merged.drop(columns="H.客数（合計）", inplace=True)

        # self.chart_cli.create_pie_chart(df=df_merged, amount_col='D.価格')
        # self.chart_cli.savefig(self.sca_s.OUTPUT_DIR, 'ABC分析_売上構成比.png')

        self.output_dict.update({'ABC分析_' + '_'.join(key_li): df_merged})

if __name__ == '__main__':
    sca = StoreCurrAnalysis()
    # store_li = ['大和乃山賊', '定楽屋', 'うおにく', 'かこい屋', 'くつろぎ屋', 'ご馳走屋名駅店', 'ご馳走屋金山店',
    #             '九州乃山賊小倉総本店', '和古屋', '楽屋', '鳥Bouno!', 'ぐるめ屋']
    # # Slice pro_plan base on process number
    # sls = len(store_li) // 3
    # store_li_sls = [store_li[sls * i:sls * (i + 1):] for i in range(3)]
    # proc = []
    # for ss in store_li_sls:
    #     proc.append(Process(target=sca.execute, args=(ss,)))
    # for p in proc:
    #     p.start()
    # for p in proc:
    #     p.join()
    sca.execute()
    print("END")
